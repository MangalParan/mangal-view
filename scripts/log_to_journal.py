#!/usr/bin/env python3
"""log_to_journal.py — parse the AI-bot trade log into Journal.xlsx.

Reads one or more bot log files (the `log.txt` style produced by the panel /
the pasted session logs) and writes a structured, reusable trading journal:

  Sheet "Trades"  — one row per ENTRY..EXIT pair (entry/exit px, PnL, R, reason,
                    SL/TP, order ids, whether it was a TP-continuation re-entry).
  Sheet "Alerts"  — every TV-ALERT webhook with its indicator fields broken out.
  Sheet "Config"  — one row per BOT START: every parameter logged ([CONFIG] line)
                    so each run's settings can be reproduced/audited.
  Sheet "Events"  — every parsed log line (timestamp, bot, category, message).
  Sheet "Summary" — per-symbol and overall stats (trades, win%, gross/net PnL).

Re-running rebuilds the workbook from the given logs, so it stays idempotent.

Usage:
    python scripts/log_to_journal.py                       # all logs/*.txt + log.txt
    python scripts/log_to_journal.py logs/trade_log_20260621.txt
    python scripts/log_to_journal.py a.txt b.txt -o Journal.xlsx
"""
import os
import re
import sys
import glob
import argparse
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LINE_RE = re.compile(r'^\[(?P<ts>[^\]]+)\]\s+(?P<rest>.*)$')

ENTRY_RE = re.compile(
    r'ENTRY (?P<side>BUY|SELL) (?P<qty>[\d.]+) (?P<sym>\S+) @ (?P<px>[\d.]+) '
    r'SL=(?P<sl>[\d.]+) TP=(?P<tp>[\d.]+) strat=(?P<strat>\S+) score=(?P<score>[\d.]+)'
    r'(?:\s+\[cap (?P<cap>[^\]]+)\])?(?:\s+[—-]\s+(?P<note>.*))?$')

EXIT_RE = re.compile(
    r'EXIT (?P<side>BUY|SELL) (?P<qty>[\d.]+) (?P<sym>\S+) @ (?P<px>[\d.]+) '
    r'\(entry (?P<entry>[\d.]+)\) PnL=(?P<pnl>\S+) reason=(?P<reason>.+?) strat=(?P<strat>\S+)$')

# Strategy Menu format (Zerodha + Delta multi-leg engine, log_options.txt):
# 'ENTRY SELL 1000 P-BTC-77000-300826 @ 63.0 role=sellPE SL=157.5 TP=44.1'
# 'EXIT BUY 1000 P-BTC-77000-300826 @ 60.0 (entry 63.0) PnL=+3.0000 reason=TP hit role=sellPE'
# Hedge legs have no SL/TP: 'SL=None TP=None'.
ENTRY2_RE = re.compile(
    r'ENTRY (?P<side>BUY|SELL) (?P<qty>[\d.]+) (?P<sym>\S+) @ (?P<px>[\d.]+) '
    r'role=(?P<role>\S+) SL=(?P<sl>None|[\d.]+) TP=(?P<tp>None|[\d.]+)$')

EXIT2_RE = re.compile(
    r'EXIT (?P<side>BUY|SELL) (?P<qty>[\d.]+) (?P<sym>\S+) @ (?P<px>[\d.]+) '
    r'\(entry (?P<entry>[\d.]+)\) PnL=(?P<pnl>\S+) reason=(?P<reason>.+?) role=(?P<role>\S+)$')

def _pnl(s):
    """Parse a PnL token tolerant of currency symbols (₹/$, or the â¹ mojibake) and
    the odd '+-0.0000' the bot prints for a flat close. Sign from any leading '-'."""
    s = (s or '').replace('+-', '-')
    neg = '-' in s
    m = re.search(r'\d+\.?\d*', s)
    if not m:
        return 0.0
    return -float(m.group()) if neg else float(m.group())

ALERT_RE = re.compile(r'\[TV-ALERT\]\s+(?P<sym>\S+)\s+(?P<sig>BUY|SELL)\s+(?P<kv>.*)$')
KV_RE = re.compile(r'(\w+)=("[^"]*"|\S+)')
ACCEPT_RE = re.compile(r'Delta accepted #(\d+)')
FAIL_RE = re.compile(r'Delta order failed: (.*)$')

# Columns broken out from TV-ALERT key=value payloads (in display order).
ALERT_FIELDS = ['price', 'exchange', 'slPct', 'tpPct', 'tf', 'rangeFilter',
                'upTrend', 'downTrend', 'halfTrend', 'bbBasis', 'bbUpper',
                'ma1', 'ma2', 'test', 'note']

HDR_FILL = PatternFill('solid', fgColor='305496')
HDR_FONT = Font(bold=True, color='FFFFFF')
WIN_FILL = PatternFill('solid', fgColor='E2EFDA')
LOSS_FILL = PatternFill('solid', fgColor='FCE4D6')


def parse_files(paths):
    events, alerts, trades = [], [], []
    # open position per bot for entry/exit pairing
    open_pos = {}
    last_event_idx_by_bot = {}

    # Gather every line from every file, drop exact duplicates (the session logs
    # overlap heavily), and sort by timestamp ascending so the journal is a single
    # clean chronological record (earliest first) regardless of file order.
    rows = []
    seen = set()
    for path in paths:
        if not os.path.exists(path):
            print('skip (missing):', path)
            continue
        with open(path, encoding='utf-8', errors='replace') as fh:
            for raw in fh:
                line = raw.rstrip('\n')
                m = LINE_RE.match(line)
                if not m or line in seen:
                    continue
                seen.add(line)
                rows.append((m.group('ts'), m.group('rest')))
    rows.sort(key=lambda r: r[0])   # ISO timestamps sort chronologically
    print('Merged {} unique log lines from {} file(s).'.format(len(rows), len(paths)))

    if True:
        if True:
            for ts, rest in rows:
                # category tags like [DELTA] [LIVE] / [TV-ALERT] / [DELTA] [STOP]
                bot, category, msg = '', '', rest
                tags = re.findall(r'^(?:\[([^\]]+)\]\s*)', rest)
                # peel leading [..] tags
                lead = re.match(r'^(\[[^\]]+\]\s*)+', rest)
                taglist = re.findall(r'\[([^\]]+)\]', lead.group(0)) if lead else []
                if taglist:
                    bot = taglist[0]
                    category = taglist[1] if len(taglist) > 1 else ''
                    msg = rest[lead.end():].strip()

                events.append({'time': ts, 'bot': bot, 'category': category, 'msg': msg, 'raw': rest})

                # --- TV alert ---
                am = ALERT_RE.search(rest)
                if am:
                    kv = dict((k, v.strip('"')) for k, v in KV_RE.findall(am.group('kv')))
                    rec = {'time': ts, 'symbol': am.group('sym'), 'signal': am.group('sig')}
                    rec.update({f: kv.get(f, '') for f in ALERT_FIELDS})
                    rec['is_test'] = 'YES' if kv.get('test') else ''
                    alerts.append(rec)
                    continue

                # --- order accept / fail attaches to the most recent trade leg ---
                acc = ACCEPT_RE.search(msg)
                fail = FAIL_RE.search(msg)
                if acc or fail:
                    pend = last_event_idx_by_bot.get(bot)
                    if pend is not None:
                        leg, which = pend
                        oid = ('#' + acc.group(1)) if acc else ('FAILED: ' + fail.group(1))
                        if which == 'entry':
                            leg['entry_order'] = (leg.get('entry_order') or oid)
                        else:
                            leg['exit_order'] = (leg.get('exit_order') or oid)
                    continue

                # --- ENTRY ---
                em = ENTRY_RE.search(msg)
                if em:
                    leg = {
                        'entry_time': ts, 'bot': bot, 'symbol': em.group('sym'),
                        'side': em.group('side'), 'qty': float(em.group('qty')),
                        'entry_px': float(em.group('px')), 'sl': float(em.group('sl')),
                        'tp': float(em.group('tp')), 'strategy': em.group('strat'),
                        'score': float(em.group('score')), 'cap': em.group('cap') or '',
                        'entry_note': (em.group('note') or '').strip(), 'role': '',
                        'exit_time': '', 'exit_px': '', 'pnl': '', 'exit_reason': '',
                        'entry_order': '', 'exit_order': '',
                    }
                    leg['continuation'] = 'YES' if 'continuation' in leg['entry_note'].lower() else ''
                    trades.append(leg)
                    open_pos[bot] = leg
                    last_event_idx_by_bot[bot] = (leg, 'entry')
                    continue

                # --- EXIT ---
                xm = EXIT_RE.search(msg)
                if xm:
                    leg = open_pos.get(bot)
                    if leg is None:  # orphan exit — make a stub
                        leg = {'entry_time': '', 'bot': bot, 'symbol': xm.group('sym'),
                               'side': 'BUY' if xm.group('side') == 'SELL' else 'SELL',
                               'qty': float(xm.group('qty')), 'entry_px': float(xm.group('entry')),
                               'sl': '', 'tp': '', 'strategy': xm.group('strat'), 'score': '',
                               'cap': '', 'entry_note': '', 'entry_order': '', 'exit_order': '',
                               'continuation': '', 'role': ''}
                        trades.append(leg)
                    leg['exit_time'] = ts
                    leg['exit_px'] = float(xm.group('px'))
                    leg['pnl'] = _pnl(xm.group('pnl'))
                    leg['exit_reason'] = xm.group('reason').strip()
                    open_pos[bot] = None
                    last_event_idx_by_bot[bot] = (leg, 'exit')
                    continue

                # --- ENTRY (Strategy Menu multi-leg format: role= identifies WHICH
                # leg, since 2+ legs — e.g. sellCE + hedgeCE — are often open at
                # once under the same bot tag, unlike the single-slot bots above) ---
                em2 = ENTRY2_RE.search(msg)
                if em2:
                    role = em2.group('role')
                    key = (bot, role)
                    sl_raw, tp_raw = em2.group('sl'), em2.group('tp')
                    leg = {
                        'entry_time': ts, 'bot': bot, 'symbol': em2.group('sym'),
                        'side': em2.group('side'), 'qty': float(em2.group('qty')),
                        'entry_px': float(em2.group('px')),
                        'sl': float(sl_raw) if sl_raw != 'None' else '',
                        'tp': float(tp_raw) if tp_raw != 'None' else '',
                        'strategy': '', 'score': '', 'cap': '', 'entry_note': '', 'role': role,
                        'exit_time': '', 'exit_px': '', 'pnl': '', 'exit_reason': '',
                        'entry_order': '', 'exit_order': '', 'continuation': '',
                    }
                    trades.append(leg)
                    open_pos[key] = leg
                    last_event_idx_by_bot[bot] = (leg, 'entry')
                    continue

                # --- EXIT (Strategy Menu multi-leg format) ---
                xm2 = EXIT2_RE.search(msg)
                if xm2:
                    role = xm2.group('role')
                    key = (bot, role)
                    leg = open_pos.get(key)
                    if leg is None:  # orphan exit — make a stub
                        leg = {'entry_time': '', 'bot': bot, 'symbol': xm2.group('sym'),
                               'side': 'BUY' if xm2.group('side') == 'SELL' else 'SELL',
                               'qty': float(xm2.group('qty')), 'entry_px': float(xm2.group('entry')),
                               'sl': '', 'tp': '', 'strategy': '', 'score': '',
                               'cap': '', 'entry_note': '', 'entry_order': '', 'exit_order': '',
                               'continuation': '', 'role': role}
                        trades.append(leg)
                    leg['exit_time'] = ts
                    leg['exit_px'] = float(xm2.group('px'))
                    leg['pnl'] = _pnl(xm2.group('pnl'))
                    leg['exit_reason'] = xm2.group('reason').strip()
                    open_pos[key] = None
                    last_event_idx_by_bot[bot] = (leg, 'exit')
                    continue

    # derived metrics
    for t in trades:
        ep, xp = t.get('entry_px'), t.get('exit_px')
        if isinstance(ep, float) and isinstance(xp, float) and ep:
            move = (xp - ep) / ep * 100.0
            t['move_pct'] = round(move if t['side'] == 'BUY' else -move, 4)
            # R multiple vs the planned SL distance
            sl = t.get('sl')
            if isinstance(sl, float) and sl:
                risk = abs(ep - sl)
                t['R'] = round((abs(xp - ep) / risk) * (1 if (t.get('pnl') or 0) >= 0 else -1), 3) if risk else ''
            else:
                t['R'] = ''
            # minutes held
            try:
                a = datetime.fromisoformat(t['entry_time']); b = datetime.fromisoformat(t['exit_time'])
                t['mins'] = round((b - a).total_seconds() / 60.0, 1)
            except Exception:
                t['mins'] = ''
        else:
            t['move_pct'] = ''; t['R'] = ''; t['mins'] = ''
    return events, alerts, trades


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = '%s1:%s%d' % (get_column_letter(1), get_column_letter(ncols), ws.max_row)


def _autofit(ws, maxw=60):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(maxw, max(8, width + 2))


def build_workbook(events, alerts, trades, out_path):
    wb = Workbook()

    # ---- Trades ----
    ws = wb.active; ws.title = 'Trades'
    cols = [('Entry Time', 'entry_time'), ('Exit Time', 'exit_time'), ('Bot', 'bot'),
            ('Role', 'role'), ('Symbol', 'symbol'), ('Side', 'side'), ('Qty', 'qty'),
            ('Entry Px', 'entry_px'), ('SL', 'sl'), ('TP', 'tp'),
            ('Exit Px', 'exit_px'), ('PnL', 'pnl'), ('Move %', 'move_pct'),
            ('R', 'R'), ('Mins', 'mins'), ('Exit Reason', 'exit_reason'),
            ('Continuation', 'continuation'), ('Strategy', 'strategy'),
            ('Score', 'score'), ('Cap/Lev', 'cap'), ('Entry Note', 'entry_note'),
            ('Entry Order', 'entry_order'), ('Exit Order', 'exit_order')]
    ws.append([h for h, _ in cols])
    for t in trades:
        ws.append([t.get(k, '') for _, k in cols])
    # color rows by PnL sign
    pnl_col = [k for _, k in cols].index('pnl') + 1
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=pnl_col).value
        if isinstance(v, (int, float)):
            fill = WIN_FILL if v >= 0 else LOSS_FILL
            for c in range(1, len(cols) + 1):
                ws.cell(row=r, column=c).fill = fill
    _style_header(ws, len(cols)); _autofit(ws)

    # ---- Alerts ----
    wa = wb.create_sheet('Alerts')
    acols = ['time', 'symbol', 'signal'] + ALERT_FIELDS
    wa.append([c.title() if c != 'is_test' else 'Test' for c in acols])
    for a in alerts:
        wa.append([a.get(c, '') for c in acols])
    _style_header(wa, len(acols)); _autofit(wa, maxw=24)

    # ---- Config (one row per BOT START — every parameter logged for that run) ----
    cfg_keys = ['claude', 'mode', 'symbol', 'symbols', 'autoSymbol', 'exchange', 'qty', 'tf',
                'capital', 'leverage', 'slPct', 'tpPct', 'maxConsec', 'maxLoss', 'maxProfit',
                'minScore', 'scoreBuffer', 'cooldownSec', 'tickSec', 'qualityFilter', 'avoidRange',
                'minER', 'rangeLook', 'maxCont', 'contSlPct', 'trendGate', 'emaMode', 'pendingSec',
                'maxSeedSlPct', 'hardStop', 'tvEnabled', 'tvSymbol', 'model', 'strategies', 'seedBias']
    cfg_rows = []
    for e in events:
        if e.get('category') == 'CONFIG':
            kv = {}
            for tok in e['msg'].split():
                if '=' in tok:
                    k, v = tok.split('=', 1); kv[k] = v
            cfg_rows.append((e['time'], e['bot'], kv))
    if cfg_rows:
        extra = []
        for _, _, kv in cfg_rows:
            for k in kv:
                if k not in cfg_keys and k not in extra:
                    extra.append(k)
        wc = wb.create_sheet('Config')
        cols = ['Time', 'Bot'] + cfg_keys + extra
        wc.append([c.title() if c in ('time', 'bot') else c for c in cols])
        for ts, bot, kv in cfg_rows:
            wc.append([ts, bot] + [kv.get(k, '') for k in cfg_keys + extra])
        _style_header(wc, len(cols)); _autofit(wc, maxw=24)

    # ---- Events (raw) ----
    we = wb.create_sheet('Events')
    we.append(['Time', 'Bot', 'Category', 'Message'])
    for e in events:
        we.append([e['time'], e['bot'], e['category'], e['msg']])
    _style_header(we, 4); _autofit(we, maxw=90)

    # ---- Summary ----
    closed = [t for t in trades if isinstance(t.get('pnl'), (int, float))]
    by_sym = {}
    for t in closed:
        by_sym.setdefault(t['symbol'], []).append(t)
    wsum = wb.create_sheet('Summary')
    wsum.append(['Symbol', 'Trades', 'Wins', 'Losses', 'Win %',
                 'Gross +', 'Gross -', 'Net PnL', 'Avg Win', 'Avg Loss',
                 'TP-cont legs', 'Reversal exits'])
    def block(name, rows):
        wins = [t for t in rows if t['pnl'] >= 0]; losses = [t for t in rows if t['pnl'] < 0]
        gp = sum(t['pnl'] for t in wins); gl = sum(t['pnl'] for t in losses)
        cont = sum(1 for t in rows if t.get('continuation') == 'YES')
        rev = sum(1 for t in rows if 'reversal' in str(t.get('exit_reason', '')).lower())
        wsum.append([name, len(rows), len(wins), len(losses),
                     round(len(wins) / len(rows) * 100, 1) if rows else 0,
                     round(gp, 4), round(gl, 4), round(gp + gl, 4),
                     round(gp / len(wins), 4) if wins else 0,
                     round(gl / len(losses), 4) if losses else 0, cont, rev])
    for sym in sorted(by_sym):
        block(sym, by_sym[sym])
    if closed:
        block('— ALL —', closed)
    _style_header(wsum, 12); _autofit(wsum)
    # bold the total row
    for c in range(1, 13):
        wsum.cell(row=wsum.max_row, column=c).font = Font(bold=True)

    # ---- Claude (the Claude AI strategy only — tracked apart from TV/algo trades,
    #      which trade very differently). Net is only summed WITHIN a symbol because
    #      Claude spans $ (Delta) and ₹ (Zerodha), so R (reward:risk realised) is the
    #      currency-agnostic headline. 'chart' flags entries whose reason cites the
    #      uploaded chart image (vision). ----
    def _rnum(x):
        try: return float(x)
        except Exception: return None
    cl = [t for t in closed if str(t.get('strategy', '')).lower() == 'claude']
    wcl = wb.create_sheet('Claude')
    tcols = [('Entry Time', 'entry_time'), ('Exit Time', 'exit_time'), ('Bot', 'bot'),
             ('Symbol', 'symbol'), ('Side', 'side'), ('Qty', 'qty'), ('Entry Px', 'entry_px'),
             ('SL', 'sl'), ('TP', 'tp'), ('Exit Px', 'exit_px'), ('PnL', 'pnl'), ('Move %', 'move_pct'),
             ('R', 'R'), ('Mins', 'mins'), ('Exit Reason', 'exit_reason'), ('Score', 'score'),
             ('Chart?', '_chart'), ('Entry Note', 'entry_note')]
    if not cl:
        wcl.append(['No Claude-strategy trades yet.'])
    else:
        wins = [t for t in cl if t['pnl'] >= 0]; losses = [t for t in cl if t['pnl'] < 0]
        Rs = [r for r in (_rnum(t.get('R')) for t in cl) if r is not None]
        charted = [t for t in cl if 'chart' in str(t.get('entry_note', '')).lower()]
        for k, v in [('Metric', 'Value'),
                     ('Claude trades (closed)', len(cl)),
                     ('Wins / Losses', '{} / {}'.format(len(wins), len(losses))),
                     ('Win %', round(len(wins) / len(cl) * 100, 1)),
                     ('Avg R (reward:risk realised)', round(sum(Rs) / len(Rs), 2) if Rs else ''),
                     ('Avg win / avg loss (per symbol ccy)', '{} / {}'.format(
                         round(sum(t['pnl'] for t in wins) / len(wins), 2) if wins else 0,
                         round(sum(t['pnl'] for t in losses) / len(losses), 2) if losses else 0)),
                     ('Trades citing the chart image', '{} of {}'.format(len(charted), len(cl)))]:
            wcl.append([k, v])
        wcl.cell(row=1, column=1).font = Font(bold=True)
        wcl.cell(row=1, column=2).font = Font(bold=True)
        wcl.append([])
        wcl.append(['Symbol', 'Trades', 'Wins', 'Win %', 'Net (own ccy)', 'Avg R'])
        _hdr2 = wcl.max_row
        bysym = {}
        for t in cl:
            bysym.setdefault(t['symbol'], []).append(t)
        for s in sorted(bysym):
            v = bysym[s]; w = [t for t in v if t['pnl'] >= 0]
            rr = [r for r in (_rnum(t.get('R')) for t in v) if r is not None]
            wcl.append([s, len(v), len(w), round(len(w) / len(v) * 100),
                        round(sum(t['pnl'] for t in v), 2), round(sum(rr) / len(rr), 2) if rr else ''])
        for c in range(1, 7):
            wcl.cell(row=_hdr2, column=c).font = Font(bold=True)
        wcl.append([])
        # full Claude trade list (newest last), color-coded by PnL
        _hdr3 = wcl.max_row + 1
        wcl.append([h for h, _ in tcols])
        for t in cl:
            row = []
            for _, k in tcols:
                if k == '_chart':
                    row.append('yes' if 'chart' in str(t.get('entry_note', '')).lower() else '')
                else:
                    row.append(t.get(k, ''))
            wcl.append(row)
        pcol = [k for _, k in tcols].index('pnl') + 1
        for r in range(_hdr3 + 1, wcl.max_row + 1):
            v = wcl.cell(row=r, column=pcol).value
            if isinstance(v, (int, float)):
                fill = WIN_FILL if v >= 0 else LOSS_FILL
                for c in range(1, len(tcols) + 1):
                    wcl.cell(row=r, column=c).fill = fill
        for c in range(1, len(tcols) + 1):
            wcl.cell(row=_hdr3, column=c).font = Font(bold=True)
    _autofit(wcl, maxw=70)

    wb.save(out_path)
    return len(trades), len(alerts), len(events), len(closed)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('logs', nargs='*', help='log files (default: logs/*.txt + log.txt)')
    ap.add_argument('-o', '--out', default='Journal.xlsx')
    args = ap.parse_args()

    paths = args.logs or (sorted(glob.glob('logs/*.txt')) + (['log.txt'] if os.path.exists('log.txt') else []))
    paths = [p for p in paths if os.path.exists(p)]
    if not paths:
        print('No log files found.'); return 1

    print('Parsing:', ', '.join(paths))
    events, alerts, trades = parse_files(paths)
    nt, na, ne, nc = build_workbook(events, alerts, trades, args.out)
    print('Wrote %s — %d trades (%d closed), %d alerts, %d events.' % (args.out, nt, nc, na, ne))
    return 0


if __name__ == '__main__':
    sys.exit(main())
